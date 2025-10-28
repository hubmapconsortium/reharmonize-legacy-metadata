#!/usr/bin/env python3
"""
Find Non-Standard Values for Data Curator Review

Analyzes processed metadata JSON files to identify values from modified metadata
that are not in the standardized value set. These non-standard values require
review by domain experts and data curators to determine if they should be added
to the standard value set or corrected.

The script uses three detection approaches:
1. Non-standard values: Checks modified_metadata against schema's standardized permissible values
2. Missing required values: Detects null or empty values in required fields per schema
3. Regex violations: Identifies values that don't match regex pattern constraints in schema

Usage:
    python find-nonstandard-values.py <input_dir> <schema_file> <output_file> <output_spreadsheet_file>

Example:
    python find-nonstandard-values.py metadata/rnaseq/output metadata/rnaseq/rnaseq-schema.json nonstandard-values.json nonstandard-values.xlsx
"""

import json
import argparse
import sys
from pathlib import Path
from collections import defaultdict
from typing import Dict, Set, Any, List, Tuple, Optional
import re

try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def load_schema(schema_file: str) -> Tuple[Dict[str, List[Any]], Set[str], Dict[str, str]]:
    """
    Load schema and extract standardized permissible values, required fields, and regex constraints.

    Args:
        schema_file: Path to the schema JSON file

    Returns:
        Tuple containing:
        - Dictionary mapping field names to their standardized permissible values (if any)
        - Set of field names marked as required
        - Dictionary mapping field names to their regex patterns (if any)
    """
    try:
        with open(schema_file, 'r', encoding='utf-8') as f:
            schema = json.load(f)
    except FileNotFoundError:
        print(f"Error: Schema file '{schema_file}' not found.", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Failed to parse schema file: {e}", file=sys.stderr)
        sys.exit(1)

    permissible_values_map = {}
    required_fields_set = set()
    regex_constraints_map = {}

    if isinstance(schema, list):
        for field in schema:
            if isinstance(field, dict) and 'name' in field:
                field_name = field['name']
                permissible_values = field.get('permissible_values')
                if permissible_values is not None:
                    permissible_values_map[field_name] = permissible_values

                # Track required fields
                if field.get('required') is True:
                    required_fields_set.add(field_name)

                # Track regex constraints
                regex_pattern = field.get('regex')
                if regex_pattern is not None and regex_pattern != "":
                    regex_constraints_map[field_name] = regex_pattern

    return permissible_values_map, required_fields_set, regex_constraints_map


def find_non_permissible_values(
    data: Dict,
    permissible_values_map: Dict[str, List[Any]]
) -> Dict[str, Set[str]]:
    """
    Find values in modified_metadata that aren't in the standardized value set.

    Args:
        data: Parsed JSON data from a processed metadata file
        permissible_values_map: Dictionary of field names to standardized permissible values

    Returns:
        Dictionary mapping field names to sets of non-standard values
    """
    non_permissible = defaultdict(set)

    if 'modified_metadata' not in data:
        return non_permissible

    modified_metadata = data['modified_metadata']

    for field_name, value in modified_metadata.items():
        # Skip null values
        if value is None:
            continue

        # Check if field has standardized permissible values in schema
        if field_name not in permissible_values_map:
            continue

        standardized_values = permissible_values_map[field_name]

        # Convert value to string for comparison (handles different types)
        value_str = str(value)

        # Check if value is not in the standardized value set
        # Need to handle both string and non-string standardized values
        is_standard = False
        for sv in standardized_values:
            if str(sv) == value_str or sv == value:
                is_standard = True
                break

        if not is_standard:
            non_permissible[field_name].add(value_str)

    return non_permissible


def find_missing_required_values(
    data: Dict,
    required_fields_set: Set[str]
) -> Dict[str, Set[str]]:
    """
    Find required fields with null or empty values in modified_metadata.

    Args:
        data: Parsed JSON data from a processed metadata file
        required_fields_set: Set of field names marked as required in schema

    Returns:
        Dictionary mapping field names to sets of problematic value indicators
    """
    missing_required = defaultdict(set)

    if 'modified_metadata' not in data:
        return missing_required

    modified_metadata = data['modified_metadata']

    for field_name in required_fields_set:
        # Check if field exists in modified_metadata
        if field_name not in modified_metadata:
            # Field completely missing - this is a problem
            missing_required[field_name].add("MISSING_FIELD")
            continue

        value = modified_metadata[field_name]

        # Check for null
        if value is None:
            missing_required[field_name].add("null")

        # Check for empty string
        elif isinstance(value, str) and value.strip() == "":
            missing_required[field_name].add("")

        # Check for empty list/dict (less common but possible)
        elif isinstance(value, list) and len(value) == 0:
            missing_required[field_name].add("[]")
        elif isinstance(value, dict) and len(value) == 0:
            missing_required[field_name].add("{}")

    return missing_required


def find_regex_violations(
    data: Dict,
    regex_constraints_map: Dict[str, str]
) -> Dict[str, Set[str]]:
    """
    Find values in modified_metadata that don't match their regex constraints.

    Args:
        data: Parsed JSON data from a processed metadata file
        regex_constraints_map: Dictionary of field names to regex patterns from schema

    Returns:
        Dictionary mapping field names to sets of values that violate regex constraints
    """
    regex_violations = defaultdict(set)

    if 'modified_metadata' not in data:
        return regex_violations

    modified_metadata = data['modified_metadata']

    for field_name, regex_pattern in regex_constraints_map.items():
        # Skip if field doesn't exist in modified_metadata
        if field_name not in modified_metadata:
            continue

        value = modified_metadata[field_name]

        # Skip null values (handled by missing_required check)
        if value is None:
            continue

        # Convert value to string for regex matching
        value_str = str(value)

        # Skip empty strings (handled by missing_required check)
        if value_str.strip() == "":
            continue

        # Try to match the regex pattern
        try:
            pattern = re.compile(regex_pattern)
            if not pattern.fullmatch(value_str):
                # Value doesn't match the regex pattern
                regex_violations[field_name].add(value_str)
        except re.error as e:
            # Invalid regex pattern in schema - log but don't crash
            print(f"Warning: Invalid regex pattern for field '{field_name}': {e}", file=sys.stderr)
            continue

    return regex_violations


def merge_results(
    non_permissible: Dict[str, Set[str]],
    missing_required: Dict[str, Set[str]],
    regex_violations: Dict[str, Set[str]]
) -> Dict[str, Set[str]]:
    """
    Merge results from all three detection approaches.

    Args:
        non_permissible: Values not in the standardized value set
        missing_required: Required fields with null or empty values
        regex_violations: Values that don't match regex constraints

    Returns:
        Merged dictionary with all non-standard values requiring review
    """
    merged = defaultdict(set)

    # Add non-permissible values
    for field_name, values in non_permissible.items():
        merged[field_name].update(values)

    # Add missing required values
    for field_name, values in missing_required.items():
        merged[field_name].update(values)

    # Add regex violations
    for field_name, values in regex_violations.items():
        merged[field_name].update(values)

    return merged


def format_output(nonstandard_values: Dict[str, Set[str]]) -> Dict[str, Any]:
    """
    Format the output: single values as strings, multiple as sorted arrays.

    Args:
        nonstandard_values: Dictionary of field names to sets of non-standard values

    Returns:
        Formatted dictionary ready for JSON output
    """
    result = {}

    for field_name, values in nonstandard_values.items():
        sorted_values = sorted(list(values))
        if len(sorted_values) == 1:
            result[field_name] = sorted_values[0]
        else:
            result[field_name] = sorted_values

    return result


def write_sheet_with_grouping(
    worksheet,
    issue_data: Dict[str, Dict[str, Set]],
    sheet_type: str
) -> None:
    """
    Write data to a worksheet with proper grouping.

    For each dataset with multiple issues, the first row shows Dataset ID and URL,
    subsequent rows have None for these columns to create visual grouping.

    Args:
        worksheet: openpyxl worksheet object
        issue_data: Dictionary mapping hubmap_id to {field_name: set([values])}
        sheet_type: Type of sheet - 'non_permissible', 'missing_required', or 'regex_violations'
    """
    # Define columns based on sheet type
    if sheet_type == 'missing_required':
        # No "Current Value" column for missing required
        columns = ['Dataset ID', 'Field Name', 'New Value', 'Dataset URL']
        has_current_value = False
    else:
        # Include "Current Value" column for non-permissible and regex violations
        columns = ['Dataset ID', 'Field Name', 'Current Value', 'New Value', 'Dataset URL']
        has_current_value = True

    # Write header row
    worksheet.append(columns)

    # Sort datasets alphabetically by hubmap_id
    sorted_datasets = sorted(issue_data.keys())

    # Write data rows
    for hubmap_id in sorted_datasets:
        field_issues = issue_data[hubmap_id]

        # Sort fields alphabetically within each dataset
        sorted_fields = sorted(field_issues.keys())

        for idx, field_name in enumerate(sorted_fields):
            values = field_issues[field_name]

            # Get the first value (should only be one per dataset+field)
            current_value = list(values)[0] if values else ""

            # First row for this dataset: include Dataset ID and URL
            # Subsequent rows: None for Dataset ID and URL
            if idx == 0:
                dataset_id = hubmap_id
                dataset_url = f"https://portal.hubmapconsortium.org/browse/{hubmap_id}"
            else:
                dataset_id = None
                dataset_url = None

            # Build row based on sheet type
            if has_current_value:
                row = [dataset_id, field_name, current_value, "", dataset_url]
            else:
                row = [dataset_id, field_name, "", dataset_url]

            worksheet.append(row)


def generate_excel_report(
    per_file_issues: Dict[str, Dict[str, Dict[str, Set]]],
    output_file: str
) -> None:
    """
    Generate Excel spreadsheet with 3 sheets for curator review.

    Args:
        per_file_issues: Dictionary with three keys (non_permissible, missing_required,
                        regex_violations), each mapping to {hubmap_id: {field_name: set([values])}}
        output_file: Path to output .xlsx file
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Cannot generate Excel report.", file=sys.stderr)
        print("Install with: pip install openpyxl", file=sys.stderr)
        return

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Create three sheets
    sheet_configs = [
        ('Non-Standard Value', per_file_issues['non_permissible'], 'non_permissible'),
        ('Missing Required Value', per_file_issues['missing_required'], 'missing_required'),
        ('Invalid Input Pattern', per_file_issues['regex_violations'], 'regex_violations')
    ]

    for sheet_name, issue_data, sheet_type in sheet_configs:
        ws = wb.create_sheet(title=sheet_name)
        write_sheet_with_grouping(ws, issue_data, sheet_type)

    # Save workbook
    try:
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
    except Exception as e:
        print(f"Error writing Excel file: {e}", file=sys.stderr)
        sys.exit(1)


def find_nonstandard_values(
    input_dir: str,
    schema_file: str,
    output_file: str,
    spreadsheet_file: str
):
    """
    Find non-standard values from modified metadata for curator review.

    Uses three detection approaches:
    1. Non-standard values in modified_metadata vs schema's standardized value set
    2. Missing required values in modified_metadata (null or empty for required fields)
    3. Regex violations in modified_metadata (values not matching regex constraints)

    Args:
        input_dir: Directory containing processed JSON files
        schema_file: Path to schema JSON file with standardized permissible values
        output_file: Path to output JSON file for non-standard values
        spreadsheet_file: Path to output Excel (.xlsx) file for spreadsheet report
    """
    input_path = Path(input_dir)

    if not input_path.exists():
        print(f"Error: Input directory '{input_dir}' not found.", file=sys.stderr)
        sys.exit(1)

    if not input_path.is_dir():
        print(f"Error: '{input_dir}' is not a directory.", file=sys.stderr)
        sys.exit(1)

    # Load schema with standardized value sets, required fields, and regex constraints
    permissible_values_map, required_fields_set, regex_constraints_map = load_schema(schema_file)

    # Aggregate non-standard values across all files (for JSON output)
    all_nonstandard_values = defaultdict(set)

    # Per-file tracking for Excel export
    per_file_issues = {
        'non_permissible': defaultdict(lambda: defaultdict(set)),
        'missing_required': defaultdict(lambda: defaultdict(set)),
        'regex_violations': defaultdict(lambda: defaultdict(set))
    }

    files_processed = 0
    non_standard_count = 0
    missing_required_count = 0
    regex_violation_count = 0

    json_files = list(input_path.glob("*.json"))

    if not json_files:
        print(f"Warning: No JSON files found in '{input_dir}'.", file=sys.stderr)

    for json_file in json_files:
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)

            # Extract hubmap_id for per-file tracking
            hubmap_id = data.get('hubmap_id', 'UNKNOWN')

            # Approach 1: Find non-standard values
            non_standard = find_non_permissible_values(data, permissible_values_map)

            # Approach 2: Find missing required values
            missing_required = find_missing_required_values(data, required_fields_set)

            # Approach 3: Find regex violations
            regex_violations = find_regex_violations(data, regex_constraints_map)

            # Merge results for this file (for JSON output)
            file_results = merge_results(non_standard, missing_required, regex_violations)

            # Aggregate to global results (for JSON output)
            for field_name, values in file_results.items():
                all_nonstandard_values[field_name].update(values)

            # Store per-file issues (for Excel output)
            for field_name, values in non_standard.items():
                per_file_issues['non_permissible'][hubmap_id][field_name].update(values)

            for field_name, values in missing_required.items():
                per_file_issues['missing_required'][hubmap_id][field_name].update(values)

            for field_name, values in regex_violations.items():
                per_file_issues['regex_violations'][hubmap_id][field_name].update(values)

            # Track counts
            if non_standard:
                non_standard_count += 1
            if missing_required:
                missing_required_count += 1
            if regex_violations:
                regex_violation_count += 1

            files_processed += 1

        except json.JSONDecodeError as e:
            print(f"Warning: Failed to parse JSON file '{json_file}': {e}", file=sys.stderr)
        except Exception as e:
            print(f"Warning: Error processing file '{json_file}': {e}", file=sys.stderr)

    # Format output
    result = format_output(all_nonstandard_values)

    # Write JSON output
    output_path = Path(output_file)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, ensure_ascii=False)
    except Exception as e:
        print(f"Error writing output file: {e}", file=sys.stderr)
        sys.exit(1)

    # Generate Excel spreadsheet
    generate_excel_report(per_file_issues, spreadsheet_file)
    print(f"Excel spreadsheet saved to: {spreadsheet_file}")

    # Print summary
    print(f"Non-standard values saved to: {output_file}")
    print(f"  Files processed: {files_processed}")
    print(f"  Files with non-standard values: {non_standard_count}")
    print(f"  Files with missing required values: {missing_required_count}")
    print(f"  Files with regex violations: {regex_violation_count}")
    total_values = sum(len(v) if isinstance(v, list) else 1 for v in result.values())
    print(f"  Total non-standard values found: {total_values}")


def main():
    """CLI entrypoint"""
    parser = argparse.ArgumentParser(
        description="Find non-standard values from modified metadata for domain expert and data curator review"
    )
    parser.add_argument(
        "input_dir",
        help="Input directory containing processed JSON files"
    )
    parser.add_argument(
        "schema_file",
        help="Path to schema JSON file with standardized permissible values"
    )
    parser.add_argument(
        "output_file",
        help="Output JSON file path for non-standard values"
    )
    parser.add_argument(
        "output_spreadsheet_file",
        help="Output Excel (.xlsx) file path for spreadsheet report"
    )

    args = parser.parse_args()

    # Validate spreadsheet file extension
    if not args.output_spreadsheet_file.endswith('.xlsx'):
        print("Error: Spreadsheet file must have .xlsx extension", file=sys.stderr)
        sys.exit(1)

    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl not installed. Cannot generate Excel report.", file=sys.stderr)
        print("Install with: pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    find_nonstandard_values(
        args.input_dir,
        args.schema_file,
        args.output_file,
        args.output_spreadsheet_file
    )


if __name__ == "__main__":
    main()
