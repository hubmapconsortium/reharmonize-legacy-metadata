#!/usr/bin/env python3
"""
TODO Excel Report Generator for Non-Standard Values

This module handles the generation of Excel spreadsheets (in todo/ folders) with
data validation dropdowns for curator review. It creates workbooks with multiple
sheets and a hidden validation sheet containing permissible values.
"""

import sys
from pathlib import Path
from typing import Dict, Set, Any, List

try:
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# Import regex hints
try:
    from utils.regex_hints import get_regex_hint
    REGEX_HINTS_AVAILABLE = True
except ImportError:
    REGEX_HINTS_AVAILABLE = False
    def get_regex_hint(field_name: str) -> str:
        """Fallback function if regex_hints module is not available."""
        return ""


def get_or_create_validation_sheet(workbook):
    """
    Get or create the hidden validation data sheet.

    Args:
        workbook: openpyxl workbook object

    Returns:
        The validation sheet (hidden)
    """
    sheet_name = '_validation_data'
    if sheet_name in workbook.sheetnames:
        return workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_state = 'hidden'
        return sheet


def write_sheet_with_grouping(
    worksheet,
    issue_data: Dict[str, Dict[str, Set]],
    sheet_type: str,
    permissible_values_map: Dict[str, List[Any]],
    validation_col_tracker: Dict[str, int]
) -> None:
    """
    Write data to a worksheet with dropdown validations.

    Dataset ID and URL appear in all rows to facilitate data filtering.

    Args:
        worksheet: openpyxl worksheet object
        issue_data: Dictionary mapping hubmap_id to {field_name: set([values])}
        sheet_type: Type of sheet - 'non_permissible', 'missing_required', or 'regex_violations'
        permissible_values_map: Dictionary mapping field names to their permissible values
        validation_col_tracker: Dictionary mapping field names to column numbers in hidden sheet
    """
    # Define columns based on sheet type
    if sheet_type == 'missing_required':
        # No "Current Value" column for missing required
        columns = ['Dataset ID', 'Field Name', 'New Value', 'Dataset URL']
        has_current_value = False
        has_regex_hint = False
    elif sheet_type == 'regex_violations':
        # Include "Current Value" and "Expected Input Hint" for regex violations
        columns = ['Dataset ID', 'Field Name', 'Current Value', 'New Value', 'Expected Input Hint', 'Dataset URL']
        has_current_value = True
        has_regex_hint = True
    else:
        # Include "Current Value" column for non-permissible
        columns = ['Dataset ID', 'Field Name', 'Current Value', 'New Value', 'Dataset URL']
        has_current_value = True
        has_regex_hint = False

    # Write header row
    worksheet.append(columns)

    # Sort datasets alphabetically by hubmap_id
    sorted_datasets = sorted(issue_data.keys())

    # Track which fields appear in which rows (for dropdown validation)
    field_to_rows = {}
    current_row = 2  # Start after header row

    # Write data rows
    for hubmap_id in sorted_datasets:
        field_issues = issue_data[hubmap_id]

        # Sort fields alphabetically within each dataset
        sorted_fields = sorted(field_issues.keys())

        for field_name in sorted_fields:
            values = field_issues[field_name]

            # Get the first value (should only be one per dataset+field)
            current_value = list(values)[0] if values else ""

            # Include Dataset ID and URL in all rows for easier filtering
            dataset_id = hubmap_id
            dataset_url = f"https://portal.hubmapconsortium.org/browse/{hubmap_id}"

            # Build row based on sheet type
            if has_regex_hint:
                # For regex violations: include Expected Input Hint column
                regex_hint = get_regex_hint(field_name)
                row = [dataset_id, field_name, current_value, "", regex_hint, dataset_url]
            elif has_current_value:
                # For non-permissible: include Current Value
                row = [dataset_id, field_name, current_value, "", dataset_url]
            else:
                # For missing required: no Current Value
                row = [dataset_id, field_name, "", dataset_url]

            worksheet.append(row)

            # Track this field's row number for dropdown validation
            if field_name not in field_to_rows:
                field_to_rows[field_name] = []
            field_to_rows[field_name].append(current_row)
            current_row += 1

    # Enable AutoFilter on the header row for easier data filtering
    if current_row > 2:  # Only add filter if there's at least one data row
        # Get the range from A1 to the last column and last row
        last_col_letter = get_column_letter(len(columns))
        filter_range = f"A1:{last_col_letter}{current_row - 1}"
        worksheet.auto_filter.ref = filter_range

    # Add dropdown data validations for fields with permissible values
    if not field_to_rows:
        # No data rows, skip validation logic
        return

    # Determine which column is "New Value"
    if has_regex_hint:
        # For regex violations: Dataset ID, Field Name, Current Value, New Value, Expected Input Hint, Dataset URL
        new_value_col_idx = 4  # Column D
    elif has_current_value:
        # For non-permissible: Dataset ID, Field Name, Current Value, New Value, Dataset URL
        new_value_col_idx = 4  # Column D
    else:
        # For missing required: Dataset ID, Field Name, New Value, Dataset URL
        new_value_col_idx = 3  # Column C

    # Create and apply data validations for fields that have entries in validation_col_tracker
    for field_name, row_numbers in field_to_rows.items():
        if field_name not in validation_col_tracker:
            continue

        col_num = validation_col_tracker[field_name]
        col_letter = get_column_letter(col_num)
        pv = permissible_values_map[field_name]

        # Create range reference to hidden sheet
        range_ref = f"_validation_data!${col_letter}$1:${col_letter}${len(pv)}"

        # Create data validation
        dv = DataValidation(type="list", formula1=range_ref, allow_blank=True)
        dv.error = 'Your entry is not in the list'
        dv.errorTitle = 'Invalid Entry'
        dv.prompt = 'Please select from the list'
        dv.promptTitle = 'List Selection'

        # Add to worksheet
        worksheet.add_data_validation(dv)

        # Apply to all rows containing this field in the "New Value" column
        new_value_col_letter = get_column_letter(new_value_col_idx)
        for row_num in row_numbers:
            dv.add(f"{new_value_col_letter}{row_num}")


def generate_excel_report(
    per_file_issues: Dict[str, Dict[str, Dict[str, Set]]],
    output_file: str,
    permissible_values_map: Dict[str, List[Any]]
) -> None:
    """
    Generate Excel spreadsheet with 3 sheets for curator review.

    Args:
        per_file_issues: Dictionary with three keys (non_permissible, missing_required,
                        regex_violations), each mapping to {hubmap_id: {field_name: set([values])}}
        output_file: Path to output .xlsx file
        permissible_values_map: Dictionary mapping field names to their permissible values
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Cannot generate Excel report.", file=sys.stderr)
        print("Install with: pip install openpyxl", file=sys.stderr)
        return

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Collect all unique fields across all issue types that have permissible values
    all_fields = set()
    for issue_type, issue_data in per_file_issues.items():
        for hubmap_id, field_issues in issue_data.items():
            for field_name in field_issues.keys():
                if field_name in permissible_values_map:
                    pv = permissible_values_map[field_name]
                    if pv and len(pv) > 0:
                        all_fields.add(field_name)

    # Create hidden sheet and populate with all permissible values
    validation_col_tracker = {}
    if all_fields:
        validation_sheet = get_or_create_validation_sheet(wb)

        # Sort fields for consistent column assignment
        sorted_fields = sorted(all_fields)
        for col_num, field_name in enumerate(sorted_fields, start=1):
            validation_col_tracker[field_name] = col_num
            permissible_values = permissible_values_map[field_name]
            for row_idx, value in enumerate(permissible_values, start=1):
                validation_sheet.cell(row=row_idx, column=col_num, value=value)

    # Create three sheets
    sheet_configs = [
        ('Non-Standard Value', per_file_issues['non_permissible'], 'non_permissible'),
        ('Missing Required Value', per_file_issues['missing_required'], 'missing_required'),
        ('Invalid Input Pattern', per_file_issues['regex_violations'], 'regex_violations')
    ]

    for sheet_name, issue_data, sheet_type in sheet_configs:
        ws = wb.create_sheet(title=sheet_name)
        write_sheet_with_grouping(ws, issue_data, sheet_type, permissible_values_map, validation_col_tracker)

    # Save workbook
    try:
        output_path = Path(output_file)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(output_file)
    except Exception as e:
        print(f"Error writing Excel file: {e}", file=sys.stderr)
        sys.exit(1)


def generate_todo_excel_reports(
    grouped_issues: Dict[str, Dict],
    output_json_path: str,
    permissible_values_map: Dict[str, List[Any]]
) -> None:
    """
    Generate TODO Excel reports grouped by group_name.

    Args:
        grouped_issues: Dictionary mapping group_slug to dict with 'issues' and 'metadata'
        output_json_path: Path to JSON output file (used to determine todo folder location)
        permissible_values_map: Dictionary mapping field names to their permissible values
    """
    if not OPENPYXL_AVAILABLE:
        print("Warning: openpyxl not installed. Cannot generate Excel reports.", file=sys.stderr)
        print("Install with: pip install openpyxl", file=sys.stderr)
        return

    # Create todo folder next to JSON output
    json_path = Path(output_json_path)
    todo_dir = json_path.parent / "todo"
    todo_dir.mkdir(parents=True, exist_ok=True)

    print(f"\nGenerating Excel reports in {todo_dir}/")

    # Generate one Excel file per group
    for group_slug, group_data in grouped_issues.items():
        per_file_issues = group_data['issues']
        group_name = group_data['group_name']
        dataset_type = group_data['dataset_type']

        # Use readable format: "Group Name (DatasetType).xlsx"
        excel_filename = f"{group_name} ({dataset_type}).xlsx"
        excel_path = todo_dir / excel_filename

        # Generate the Excel report
        generate_excel_report(per_file_issues, str(excel_path), permissible_values_map)
        print(f"  - {excel_filename}")
